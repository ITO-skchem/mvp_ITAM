import json
import re
from io import BytesIO
from urllib.parse import urlencode

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from ai_search.services import AssetSearchService
from assets.models import InfraAsset
from core.models import AuditLog, Code
from masters.models import (
    Component,
    ComponentAlias,
    ConfigurationComponentMapping,
    ConfigurationMaster,
    PersonMaster,
    ServiceAttribute,
    ServiceConfigurationMapping,
    ServiceMaster,
)
from masters.service_person_grid import SERVICE_PERSON_GRID_COLUMNS, attribute_codes_for_grid


SEARCH_TERM_PATTERN = re.compile(r"([^\s:][^:]*?)\s*:\s*([^\s]+)")
SERVICE_SEARCH_FIELD_ALIASES = {
    "서비스명": "name",
    "서비스": "name",
    "name": "name",
    "분류": "category_code",
    "category": "category_code",
    "구축": "build_type_code",
    "build": "build_type_code",
    "itgc": "itgc_code",
    "등급": "service_grade_code",
    "grade": "service_grade_code",
    "상태": "status_code",
    "status": "status_code",
}
SERVICE_ROLE_FIELD_ALIASES = {
    "dt팀": "SVC_ATTR_PERSON_DT_TEAM",
    "dt": "SVC_ATTR_PERSON_DT_TEAM",
    "관리자": "SVC_ATTR_PERSON_ADMIN",
    "admin": "SVC_ATTR_PERSON_ADMIN",
    "운영자": "SVC_ATTR_PERSON_OPERATOR",
    "operator": "SVC_ATTR_PERSON_OPERATOR",
    "infra담당자": "SVC_ATTR_PERSON_INFRA_OPERATOR",
    "infra담당": "SVC_ATTR_PERSON_INFRA_OPERATOR",
    "infra": "SVC_ATTR_PERSON_INFRA_OPERATOR",
}
SERVICE_CODE_FIELD_MAP = {
    "category_code": ("service_category", "category_code"),
    "build_type_code": ("build_type", "build_type_code"),
    "itgc_code": ("yn_flag", "itgc_code"),
    "service_grade_code": ("service_grade", "service_grade_code"),
    "status_code": ("service_status", "status_code"),
}
PERSON_SEARCH_FIELD_ALIASES = {
    "성명": "name",
    "이름": "name",
    "name": "name",
    "사번": "employee_no",
    "employee": "employee_no",
    "회사명": "company",
    "회사": "company",
    "역할": "role_code",
    "상주여부": "resident_type_code",
    "소속": "affiliation_code",
    "성별": "gender_code",
    "상태": "status_code",
    "담당업무": "assigned_service_names",
    "업무": "assigned_service_names",
}
PERSON_CODE_FIELD_MAP = {
    "role_code": ("person_role", "role_code"),
    "resident_type_code": ("resident_type", "resident_type_code"),
    "affiliation_code": ("affiliation", "affiliation_code"),
    "gender_code": ("gender", "gender_code"),
    "status_code": ("person_status", "status_code"),
}
CONFIG_SEARCH_FIELD_ALIASES = {
    "구성명": "hostname",
    "hostname": "hostname",
    "ip": "ip",
    "url": "url",
    "구성유형": "server_type_code",
    "운영개발": "operation_dev_code",
    "운영/개발": "operation_dev_code",
    "인프라구분": "infra_type_code",
    "위치": "location_code",
    "네트워크": "network_zone_code",
    "연결서비스": "__connected_service__",
}
CONFIG_CONNECTED_SERVICE_KEY = "__connected_service__"
CONFIG_CODE_FIELD_MAP = {
    "server_type_code": ("config_type", "server_type_code"),
    "operation_dev_code": ("operation_type", "operation_dev_code"),
    "infra_type_code": ("infra_type", "infra_type_code"),
    "location_code": ("infra_location", "location_code"),
    "network_zone_code": ("network_zone", "network_zone_code"),
}
COMPONENT_SEARCH_FIELD_ALIASES = {
    "컴포넌트명": "product_name",
    "제품명": "product_name",
    "product": "product_name",
    "버전": "version",
    "version": "version",
    "벤더명": "vendor_name",
    "벤더": "vendor_name",
    "vendor": "vendor_name",
    "cpe": "cpe_name",
    "유형": "component_type_code",
    "지원여부": "support_status_code",
}
COMPONENT_CODE_FIELD_MAP = {
    "component_type_code": ("component_type", "component_type_code"),
    "support_status_code": ("support_status", "support_status_code"),
}


def split_csv_tokens(raw):
    return [t.strip() for t in str(raw or "").split(",") if t.strip()]


def join_csv_tokens(values):
    return ", ".join(v for v in values if str(v or "").strip())


def sync_service_configuration_mappings(configuration: ConfigurationMaster, connected_service_raw: str) -> None:
    """구성에 대한 서비스-구성 매핑을 입력값(쉼표 구분 서비스명)에 맞게 교체한다."""
    ServiceConfigurationMapping.objects.filter(configuration=configuration).delete()
    seen = set()
    for token in split_csv_tokens(connected_service_raw):
        svc = ServiceMaster.objects.filter(name=token).first()
        if svc and svc.pk not in seen:
            seen.add(svc.pk)
            ServiceConfigurationMapping.objects.create(service=svc, configuration=configuration)


def sync_configuration_component_mappings(configuration: ConfigurationMaster, row: dict, type_codes: list[str]) -> None:
    """구성-컴포넌트 매핑을 그리드 행의 유형별 컴포넌트 ID CSV(`component_<type>` 키)로부터 교체한다.

    각 셀 값은 쉼표 구분된 컴포넌트 PK 문자열이며, 유형이 다른 컴포넌트가 섞여 들어와도
    실제 컴포넌트의 `component_type_code` 값을 신뢰해 저장한다.
    """
    desired_ids: list[int] = []
    for tc in type_codes:
        raw = row.get(f"component_{tc}", "")
        for token in split_csv_tokens(raw):
            if token.isdigit():
                desired_ids.append(int(token))
    seen = set()
    unique_ids = []
    for cid in desired_ids:
        if cid in seen:
            continue
        seen.add(cid)
        unique_ids.append(cid)
    valid_ids = set(Component.objects.filter(pk__in=unique_ids).values_list("pk", flat=True))
    ConfigurationComponentMapping.objects.filter(configuration=configuration).delete()
    for cid in unique_ids:
        if cid in valid_ids:
            ConfigurationComponentMapping.objects.create(configuration=configuration, component_id=cid)


def parse_person_ids_from_attr_value(raw):
    ids = []
    for token in split_csv_tokens(raw):
        if token.isdigit():
            ids.append(int(token))
    return ids


def normalize_service_search_key(raw_key: str) -> str:
    return re.sub(r"\s+", "", str(raw_key or "").strip().lower())


def parse_structured_terms(query: str):
    terms = []
    spans = []
    for m in SEARCH_TERM_PATTERN.finditer(query or ""):
        key = normalize_service_search_key(m.group(1))
        value = (m.group(2) or "").strip()
        if key and value:
            terms.append((key, value))
            spans.append(m.span())
    if not spans:
        return [], (query or "").strip()
    rest = query or ""
    for s, e in reversed(spans):
        rest = rest[:s] + " " + rest[e:]
    return terms, " ".join(rest.split())


def matched_service_ids_for_person_keyword(keyword: str, acode: str | None = None):
    kw = (keyword or "").strip()
    if not kw:
        return set()
    person_ids = set(
        PersonMaster.objects.filter(Q(name__icontains=kw) | Q(employee_no__icontains=kw)).values_list("id", flat=True)
    )
    if not person_ids:
        return set()
    attr_qs = ServiceAttribute.objects.filter(attribute_code_id__in=attribute_codes_for_grid())
    if acode:
        attr_qs = attr_qs.filter(attribute_code_id=acode)
    pre = Q()
    for pid in person_ids:
        pre |= Q(value__contains=str(pid))
    if pre:
        attr_qs = attr_qs.filter(pre)
    matched_service_ids = set()
    for service_id, raw in attr_qs.values_list("service_id", "value"):
        if set(parse_person_ids_from_attr_value(raw)) & person_ids:
            matched_service_ids.add(service_id)
    return matched_service_ids


def q_for_service_structured_term(key: str, value: str):
    field_key = SERVICE_SEARCH_FIELD_ALIASES.get(key)
    if field_key in SERVICE_CODE_FIELD_MAP:
        group_key, model_field = SERVICE_CODE_FIELD_MAP[field_key]
        code_ids = list(
            Code.objects.filter(group__key=group_key, group__is_active=True, is_active=True)
            .filter(Q(code__icontains=value) | Q(name__icontains=value))
            .values_list("id", flat=True)
        )
        return Q(pk__in=[]) if not code_ids else Q(**{f"{model_field}_id__in": code_ids})
    if field_key == "name":
        return Q(name__icontains=value)

    acode = SERVICE_ROLE_FIELD_ALIASES.get(key)
    if acode:
        matched_ids = matched_service_ids_for_person_keyword(value, acode=acode)
        return Q(pk__in=matched_ids) if matched_ids else Q(pk__in=[])
    return None


def q_for_service_code_fields_keyword(keyword: str):
    kw = (keyword or "").strip()
    if not kw:
        return Q()
    q = Q()
    for group_key, model_field in SERVICE_CODE_FIELD_MAP.values():
        code_ids = list(
            Code.objects.filter(group__key=group_key, group__is_active=True, is_active=True)
            .filter(Q(code__icontains=kw) | Q(name__icontains=kw))
            .values_list("id", flat=True)
        )
        if code_ids:
            q |= Q(**{f"{model_field}_id__in": code_ids})
    return q


def q_for_code_fields_keyword(keyword: str, code_field_map: dict[str, tuple[str, str]]):
    kw = (keyword or "").strip()
    if not kw:
        return Q()
    q = Q()
    for group_key, model_field in code_field_map.values():
        code_ids = list(
            Code.objects.filter(group__key=group_key, group__is_active=True, is_active=True)
            .filter(Q(code__icontains=kw) | Q(name__icontains=kw))
            .values_list("id", flat=True)
        )
        if code_ids:
            q |= Q(**{f"{model_field}_id__in": code_ids})
    return q


def matched_person_ids_for_service_keyword(keyword: str):
    kw = (keyword or "").strip()
    if not kw:
        return set()
    target_service_ids = set(
        ServiceMaster.objects.filter(Q(name__icontains=kw) | Q(service_mgmt_no__icontains=kw)).values_list("id", flat=True)
    )
    if not target_service_ids:
        return set()
    matched_person_ids = set()
    for svc_id, raw in ServiceAttribute.objects.filter(attribute_code_id__in=attribute_codes_for_grid()).values_list("service_id", "value"):
        if svc_id in target_service_ids:
            matched_person_ids.update(parse_person_ids_from_attr_value(raw))
    return matched_person_ids


def person_option_map_by_role():
    from collections import defaultdict

    grouped = defaultdict(list)
    role_codes_needed = [c["role_code"] for c in SERVICE_PERSON_GRID_COLUMNS]
    qs = PersonMaster.objects.select_related("role_code").order_by("name", "employee_no")
    for p in qs:
        rc = getattr(p.role_code, "code", None) if p.role_code else None
        if rc not in role_codes_needed:
            continue
        emp = (p.employee_no or "").strip()
        grouped[rc].append({"id": p.pk, "label": f"{p.name} ({emp})" if emp else p.name})
    return grouped


def service_person_columns_with_options():
    grouped = person_option_map_by_role()
    return [{**col, "options": list(grouped[col["role_code"]])} for col in SERVICE_PERSON_GRID_COLUMNS]


def person_label_map():
    return {p.pk: (p.name or "").strip() for p in PersonMaster.objects.all().order_by("name", "employee_no")}


def service_lookup_maps():
    by_name = {}
    by_mgmt = {}
    for s in ServiceMaster.objects.all().order_by("service_mgmt_no"):
        by_name[s.name] = s
        by_mgmt[s.service_mgmt_no] = s
    return by_name, by_mgmt


def set_service_attr_person_ids(service_obj, acode, person_ids):
    clean_ids = sorted({int(x) for x in person_ids if str(x).isdigit() or isinstance(x, int)})
    if clean_ids:
        ServiceAttribute.objects.update_or_create(
            service=service_obj,
            attribute_code_id=acode,
            defaults={"value": join_csv_tokens([str(i) for i in clean_ids])},
        )
    else:
        ServiceAttribute.objects.filter(service=service_obj, attribute_code_id=acode).delete()


def sync_service_person_attributes_from_service_grid_row(service_obj, row):
    p_qs = PersonMaster.objects.all().order_by("name", "employee_no")
    label_to_id = {f"{p.name}({p.employee_no})": p.pk for p in p_qs}
    name_to_id = {p.name: p.pk for p in p_qs}
    emp_to_id = {p.employee_no: p.pk for p in p_qs}
    for col in SERVICE_PERSON_GRID_COLUMNS:
        acode = col["attribute_code"]
        key = f"attr_{acode}"
        parsed_ids = []
        for token in split_csv_tokens(row.get(key)):
            if token.isdigit() and PersonMaster.objects.filter(pk=int(token)).exists():
                parsed_ids.append(int(token))
            elif token in label_to_id:
                parsed_ids.append(label_to_id[token])
            elif token in name_to_id:
                parsed_ids.append(name_to_id[token])
            elif token in emp_to_id:
                parsed_ids.append(emp_to_id[token])
        set_service_attr_person_ids(service_obj, acode, parsed_ids)


def sync_person_service_assignments_from_row(person_obj, row):
    by_name, by_mgmt = service_lookup_maps()
    role_to_acode = {c["role_code"]: c["attribute_code"] for c in SERVICE_PERSON_GRID_COLUMNS}
    target_role = getattr(person_obj.role_code, "code", "") if person_obj.role_code else ""
    target_acode = role_to_acode.get(target_role)

    selected_service_ids = set()
    for token in split_csv_tokens(row.get("assigned_service_names")):
        svc = by_name.get(token) or by_mgmt.get(token)
        if svc:
            selected_service_ids.add(svc.pk)

    all_services = list(ServiceMaster.objects.all())
    all_acodes = list(role_to_acode.values())

    # 역할 변경/서비스 해제 케이스 포함: 먼저 모든 역할 속성에서 현재 담당자 제거
    for svc in all_services:
        for acode in all_acodes:
            curr = ServiceAttribute.objects.filter(service=svc, attribute_code_id=acode).first()
            curr_set = set(parse_person_ids_from_attr_value(curr.value if curr else ""))
            if person_obj.pk in curr_set:
                curr_set.discard(person_obj.pk)
                set_service_attr_person_ids(svc, acode, sorted(curr_set))

    # 현재 선택 역할의 서비스 목록에만 다시 추가
    if target_acode:
        for svc in all_services:
            if svc.pk not in selected_service_ids:
                continue
            curr = ServiceAttribute.objects.filter(service=svc, attribute_code_id=target_acode).first()
            curr_set = set(parse_person_ids_from_attr_value(curr.value if curr else ""))
            curr_set.add(person_obj.pk)
            set_service_attr_person_ids(svc, target_acode, sorted(curr_set))


def parse_date(val):
    if val in (None, ""):
        return None
    dt = pd.to_datetime(val, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.date()


def to_excel_response(df, filename):
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    content = out.getvalue()
    res = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    res["Content-Disposition"] = f'attachment; filename="{filename}"'
    return res


def to_excel_multi_response(sheet_map, filename, readonly_columns_map=None, code_columns_map=None):
    readonly_columns_map = readonly_columns_map or {}
    code_columns_map = code_columns_map or {}
    readonly_fill = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
    code_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet_name, df in sheet_map.items():
            ws_name = sheet_name[:31]
            df.to_excel(writer, index=False, sheet_name=ws_name)
            ws = writer.book[ws_name]
            code_cols = set(code_columns_map.get(sheet_name, []))
            readonly_cols = set(readonly_columns_map.get(sheet_name, []))
            for idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(idx)
                if col_name in code_cols:
                    for row_idx in range(1, ws.max_row + 1):
                        ws[f"{col_letter}{row_idx}"].fill = code_fill
                if col_name not in readonly_cols:
                    continue
                for row_idx in range(1, ws.max_row + 1):
                    ws[f"{col_letter}{row_idx}"].fill = readonly_fill
    content = out.getvalue()
    res = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    res["Content-Disposition"] = f'attachment; filename="{filename}"'
    return res


def code_reference_df(group_keys):
    rows = []
    for gk in group_keys:
        for code, name in (
            Code.objects.filter(group__key=gk, group__is_active=True, is_active=True)
            .order_by("sort_order", "code")
            .values_list("code", "name")
        ):
            rows.append({"코드그룹": gk, "코드값": code, "코드명": name})
    return pd.DataFrame(rows, columns=["코드그룹", "코드값", "코드명"])


def build_model_text_search_q(model_class, query):
    text = (query or "").strip()
    if not text:
        return Q()
    q = Q()
    for field in model_class._meta.fields:
        if field.get_internal_type() in {
            "CharField",
            "TextField",
            "EmailField",
            "URLField",
            "GenericIPAddressField",
        }:
            q |= Q(**{f"{field.name}__icontains": text})
    return q


def build_list_redirect(path, query="", page="", failed_ids=None):
    params = {}
    if query:
        params["q"] = query
    if page:
        params["page"] = page
    if failed_ids:
        params["failed_ids"] = ",".join(str(x) for x in failed_ids)
    return f"{path}?{urlencode(params)}" if params else path


def paginate_all_rows(qs, page_param):
    """목록 그리드에서 전체 행을 표시한다(고정 100건 페이징 없음)."""
    total = qs.count()
    paginator = Paginator(qs, max(total, 1))
    return paginator.get_page(page_param)


def code_values(group_key):
    return list(
        Code.objects.filter(group__key=group_key, group__is_active=True, is_active=True)
        .order_by("sort_order", "code")
        .values_list("code", flat=True)
    )


def code_choice_options(group_key):
    """그리드 셀렉트용: value는 code, 라벨은 Code.name(표시명)."""
    return list(
        Code.objects.filter(group__key=group_key, group__is_active=True, is_active=True)
        .order_by("sort_order", "code")
        .values("code", "name")
    )


# 담당자 역할 코드 개명 이전 코드값 → 현재 코드값 (엑셀 재업로드 등)
LEGACY_PERSON_ROLE_CODES = {"APPL_OPS": "OPERATOR", "INFRA_OPS": "INFRA_OPERATOR"}
# 컴포넌트 유형 코드 재정리 이전 값 → 현재 코드값
LEGACY_COMPONENT_TYPE_CODES = {"APP": "기타", "MIDDLEWARE": "middleware", "OS": "OS", "DB": "DB"}
LEGACY_SERVICE_STATUS_CODES = {"ACTIVE": "운영중", "active": "운영중", "PAUSED": "대기", "paused": "대기", "END": "종료", "end": "종료"}
LEGACY_BUILD_TYPE_CODES = {"NEW": "SI개발", "RENEWAL": "SI개발", "MAINT": "솔루션"}
LEGACY_PERSON_STATUS_CODES = {"ACTIVE": "투입", "LEAVE": "대기", "END": "종료", "재직": "투입", "휴직": "대기"}
LEGACY_YN_FLAG_CODES = {"Y": "예", "N": "아니오"}
LEGACY_OPERATION_TYPE_CODES = {"OPS": "운영", "DEV": "개발", "BOTH": "운영", "운영/개발": "운영"}
LEGACY_CONFIG_TYPE_CODES = {"ETC": "기타", "CACHE": "기타"}
LEGACY_INFRA_TYPE_CODES = {"ONPREM": "OnPrem", "온프레미스": "OnPrem", "AZURE": "기타", "GCP": "기타"}
LEGACY_INFRA_LOCATION_CODES = {"DC1": "판교", "DC2": "청주", "센터1": "판교", "센터2": "청주"}
LEGACY_NETWORK_ZONE_CODES = {"INTERNAL": "내부망", "EXTERNAL": "외부망", "내부": "내부망", "외부": "외부망"}


def code_from_value(group_key, value):
    v = (value or "").strip()
    if not v:
        return None
    if group_key == "person_role" and v in LEGACY_PERSON_ROLE_CODES:
        v = LEGACY_PERSON_ROLE_CODES[v]
    if group_key == "service_status" and v in LEGACY_SERVICE_STATUS_CODES:
        v = LEGACY_SERVICE_STATUS_CODES[v]
    if group_key == "build_type" and v in LEGACY_BUILD_TYPE_CODES:
        v = LEGACY_BUILD_TYPE_CODES[v]
    if group_key == "person_status" and v in LEGACY_PERSON_STATUS_CODES:
        v = LEGACY_PERSON_STATUS_CODES[v]
    if group_key == "yn_flag" and v in LEGACY_YN_FLAG_CODES:
        v = LEGACY_YN_FLAG_CODES[v]
    if group_key == "component_type" and v in LEGACY_COMPONENT_TYPE_CODES:
        v = LEGACY_COMPONENT_TYPE_CODES[v]
    if group_key == "operation_type" and v in LEGACY_OPERATION_TYPE_CODES:
        v = LEGACY_OPERATION_TYPE_CODES[v]
    if group_key == "config_type" and v in LEGACY_CONFIG_TYPE_CODES:
        v = LEGACY_CONFIG_TYPE_CODES[v]
    if group_key == "infra_type" and v in LEGACY_INFRA_TYPE_CODES:
        v = LEGACY_INFRA_TYPE_CODES[v]
    if group_key == "infra_location" and v in LEGACY_INFRA_LOCATION_CODES:
        v = LEGACY_INFRA_LOCATION_CODES[v]
    if group_key == "network_zone" and v in LEGACY_NETWORK_ZONE_CODES:
        v = LEGACY_NETWORK_ZONE_CODES[v]
    base_qs = Code.objects.filter(group__key=group_key, group__is_active=True, is_active=True)
    by_code = base_qs.filter(code=v).first()
    if by_code:
        return by_code
    # 엑셀·화면은 Code.name(표시명)으로 내려받는 경우 동일하게 FK로 해석
    by_name = base_qs.filter(name=v).first()
    if by_name:
        return by_name
    return base_qs.filter(name__iexact=v).first()


def audit_actor_label(request):
    user = getattr(request, "user", None)
    if user and getattr(user, "is_authenticated", False):
        name = getattr(user, "username", "") or getattr(user, "get_username", lambda: "")()
        return (name or str(getattr(user, "pk", ""))).strip() or "unknown"
    return ""


@login_required
def dashboard(request):
    context = {
        "counts": {
            "infra": InfraAsset.objects.count(),
            "service": ServiceMaster.objects.count(),
            "person": PersonMaster.objects.count(),
            "component": Component.objects.count(),
            "configuration": ConfigurationMaster.objects.count(),
            "audit": AuditLog.objects.count(),
        }
    }
    return render(request, "web/dashboard.html", context)


@login_required
@permission_required("assets.view_infraasset", raise_exception=True)
def asset_list(request):
    # 과거(미사용) 화면: 현재는 시스템 통합정보 통합 View(`integrated_view`)를 사용한다.
    return redirect("web:integrated_view")


@login_required
@permission_required("assets.view_infraasset", raise_exception=True)
def asset_detail(request, pk):
    # 과거(미사용) 화면: 현재는 시스템 통합정보 통합 View(`integrated_view`)를 사용한다.
    return redirect("web:integrated_view")


@login_required
@permission_required("masters.view_servicemaster", raise_exception=True)
def service_master_list(request):
    query = request.GET.get("q", "").strip()
    page = request.GET.get("page", "").strip()
    attr_codes = attribute_codes_for_grid()
    qs = (
        ServiceMaster.objects.select_related(
            "category_code", "status_code", "build_type_code", "itgc_code", "service_grade_code"
        )
        .prefetch_related(
            Prefetch(
                "service_attributes",
                queryset=ServiceAttribute.objects.filter(attribute_code_id__in=attr_codes),
            )
        )
        .order_by("service_mgmt_no")
    )
    if query:
        terms, free_text = parse_structured_terms(query)
        if terms:
            for key, value in terms:
                term_q = q_for_service_structured_term(key, value)
                if term_q is None:
                    continue
                qs = qs.filter(term_q)
            if free_text:
                free_q = build_model_text_search_q(ServiceMaster, free_text)
                free_q |= q_for_service_code_fields_keyword(free_text)
                matched_ids = matched_service_ids_for_person_keyword(free_text)
                if matched_ids:
                    free_q |= Q(id__in=matched_ids)
                qs = qs.filter(free_q)
            qs = qs.distinct()
        else:
            base_q = build_model_text_search_q(ServiceMaster, query)
            base_q |= q_for_service_code_fields_keyword(query)
            matched_ids = matched_service_ids_for_person_keyword(query)
            if matched_ids:
                base_q |= Q(id__in=matched_ids)
            qs = qs.filter(base_q).distinct()

    if request.GET.get("export") == "1":
        p_label_map = person_label_map()
        rows = []
        for s in qs:
            role_map = {}
            for sa in s.service_attributes.all():
                role_map[sa.attribute_code_id] = join_csv_tokens(
                    [p_label_map.get(pid, str(pid)) for pid in parse_person_ids_from_attr_value(sa.value)]
                )
            row = {
                "서비스ID": s.service_mgmt_no,
                "서비스명": s.name,
                "분류": getattr(s.category_code, "name", "") if s.category_code else "",
            }
            for col in SERVICE_PERSON_GRID_COLUMNS:
                row[col["label"]] = role_map.get(col["attribute_code"], "")
            row.update(
                {
                    "구축": getattr(s.build_type_code, "name", "") if s.build_type_code else "",
                    "ITGC": getattr(s.itgc_code, "name", "") if s.itgc_code else "",
                    "등급": getattr(s.service_grade_code, "name", "") if s.service_grade_code else "",
                    "상태": getattr(s.status_code, "name", "") if s.status_code else "",
                    "오픈일": s.opened_at,
                    "종료일": s.ended_at,
                    "설명": s.description,
                }
            )
            rows.append(row)
        main_df = pd.DataFrame(rows)
        ref_df = code_reference_df(["service_category", "service_status", "build_type", "yn_flag", "service_grade"])
        return to_excel_multi_response(
            {"서비스마스터": main_df, "코드참조": ref_df},
            "service_master.xlsx",
            readonly_columns_map={"서비스마스터": [c["label"] for c in SERVICE_PERSON_GRID_COLUMNS]},
            code_columns_map={"서비스마스터": ["분류", "구축", "ITGC", "등급", "상태"]},
        )

    if request.method == "POST" and request.POST.get("action") == "import":
        up = request.FILES.get("excel_file")
        if up:
            df = pd.read_excel(up)
            actor = audit_actor_label(request)
            for _, rec in df.fillna("").iterrows():
                payload = {
                    "name": str(rec.get("서비스명", "")).strip(),
                    "category_code": code_from_value("service_category", rec.get("분류")),
                    "status_code": code_from_value("service_status", rec.get("상태")),
                    "build_type_code": code_from_value("build_type", rec.get("구축")),
                    "itgc_code": code_from_value("yn_flag", rec.get("ITGC")),
                    "service_grade_code": code_from_value("service_grade", rec.get("등급")),
                    "opened_at": parse_date(rec.get("오픈일")),
                    "ended_at": parse_date(rec.get("종료일")),
                    "description": str(rec.get("설명", "")).strip(),
                }
                svc_id = str(rec.get("서비스ID", "")).strip()
                obj = ServiceMaster.objects.filter(service_mgmt_no=svc_id).first() if svc_id else None
                if obj:
                    for k, v in payload.items():
                        setattr(obj, k, v)
                    obj.updated_by = actor
                    obj.save()
                elif payload["name"]:
                    ServiceMaster.objects.create(**payload, created_by=actor, updated_by=actor)
            messages.success(request, "서비스 마스터 엑셀 업로드 완료")
        return redirect(build_list_redirect(request.path, query=query, page=page))

    if request.method == "POST" and request.POST.get("action") == "save":
        rows = json.loads(request.POST.get("rows_json", "[]"))
        deleted_ids = json.loads(request.POST.get("deleted_ids_json", "[]"))
        actor = audit_actor_label(request)
        ServiceMaster.objects.filter(pk__in=deleted_ids).delete()
        for row in rows:
            pk = str(row.get("id", "")).strip()
            payload = {
                "name": (row.get("name") or "").strip(),
                "category_code": code_from_value("service_category", row.get("category_code")),
                "status_code": code_from_value("service_status", row.get("status_code")),
                "build_type_code": code_from_value("build_type", row.get("build_type_code")),
                "itgc_code": code_from_value("yn_flag", row.get("itgc_code")),
                "service_grade_code": code_from_value("service_grade", row.get("service_grade_code")),
                "opened_at": parse_date(row.get("opened_at")),
                "ended_at": parse_date(row.get("ended_at")),
                "description": (row.get("description") or "").strip(),
            }
            obj = None
            if pk:
                obj = ServiceMaster.objects.get(pk=pk)
                for k, v in payload.items():
                    setattr(obj, k, v)
                obj.updated_by = actor
                obj.save()
            elif payload["name"]:
                obj = ServiceMaster.objects.create(**payload, created_by=actor, updated_by=actor)
        messages.success(request, "서비스 마스터 저장 완료")
        return redirect(build_list_redirect(request.path, query=query, page=page))

    page_obj = paginate_all_rows(qs, request.GET.get("page"))
    p_label_map = person_label_map()
    for s in page_obj.object_list:
        raw_sa = {}
        for sa in s.service_attributes.all():
            raw_sa[sa.attribute_code_id] = str(sa.value or "").strip()
        s.person_slot_values = {
            col["attribute_code"]: join_csv_tokens(
                [
                    p_label_map.get(pid, str(pid))
                    for pid in parse_person_ids_from_attr_value(raw_sa.get(col["attribute_code"], ""))
                ]
            )
            for col in SERVICE_PERSON_GRID_COLUMNS
        }
    return render(
        request,
        "web/service_master_list.html",
        {
            "items": page_obj.object_list,
            "page_obj": page_obj,
            "q": query,
            "service_category_options": code_choice_options("service_category"),
            "service_status_options": code_choice_options("service_status"),
            "build_type_options": code_choice_options("build_type"),
            "yn_flag_options": code_choice_options("yn_flag"),
            "service_grade_options": code_choice_options("service_grade"),
            "service_person_columns": service_person_columns_with_options(),
        },
    )


@login_required
@permission_required("masters.view_personmaster", raise_exception=True)
def person_master_list(request):
    query = request.GET.get("q", "").strip()
    page = request.GET.get("page", "").strip()
    qs = PersonMaster.objects.select_related(
        "role_code", "resident_type_code", "affiliation_code", "gender_code", "status_code"
    ).order_by("person_mgmt_no")
    if query:
        terms, free_text = parse_structured_terms(query)
        if terms:
            for key, value in terms:
                field_key = PERSON_SEARCH_FIELD_ALIASES.get(key)
                if field_key in PERSON_CODE_FIELD_MAP:
                    term_q = q_for_code_fields_keyword(value, {field_key: PERSON_CODE_FIELD_MAP[field_key]})
                    qs = qs.filter(term_q if term_q else Q(pk__in=[]))
                elif field_key == "assigned_service_names":
                    ids = matched_person_ids_for_service_keyword(value)
                    qs = qs.filter(Q(id__in=ids) if ids else Q(pk__in=[]))
                elif field_key in {"name", "employee_no", "company"}:
                    qs = qs.filter(Q(**{f"{field_key}__icontains": value}))
            if free_text:
                free_q = build_model_text_search_q(PersonMaster, free_text)
                free_q |= q_for_code_fields_keyword(free_text, PERSON_CODE_FIELD_MAP)
                person_ids = matched_person_ids_for_service_keyword(free_text)
                if person_ids:
                    free_q |= Q(id__in=person_ids)
                qs = qs.filter(free_q)
            qs = qs.distinct()
        else:
            base_q = build_model_text_search_q(PersonMaster, query)
            base_q |= q_for_code_fields_keyword(query, PERSON_CODE_FIELD_MAP)
            person_ids = matched_person_ids_for_service_keyword(query)
            if person_ids:
                base_q |= Q(id__in=person_ids)
            qs = qs.filter(base_q).distinct()

    if request.GET.get("export") == "1":
        role_to_acode = {c["role_code"]: c["attribute_code"] for c in SERVICE_PERSON_GRID_COLUMNS}
        sa_lookup = {
            (sa.service_id, sa.attribute_code_id): set(parse_person_ids_from_attr_value(sa.value))
            for sa in ServiceAttribute.objects.filter(attribute_code_id__in=attribute_codes_for_grid())
        }
        services = list(ServiceMaster.objects.all().order_by("service_mgmt_no"))
        rows = []
        for p in qs:
            my_acode = role_to_acode.get(getattr(p.role_code, "code", "") if p.role_code else "")
            names = [svc.name for svc in services if my_acode and p.pk in sa_lookup.get((svc.pk, my_acode), set())]
            rows.append(
                {
                    "담당자ID": p.person_mgmt_no,
                    "성명": p.name,
                    "역할": getattr(p.role_code, "name", "") if p.role_code else "",
                    "담당업무": join_csv_tokens(names),
                    "상주여부": getattr(p.resident_type_code, "name", "") if p.resident_type_code else "",
                    "소속": getattr(p.affiliation_code, "name", "") if p.affiliation_code else "",
                    "회사명": p.company,
                    "전화번호": p.phone,
                    "내부메일": p.email,
                    "외부메일": p.external_email,
                    "사번": p.employee_no,
                    "성별": getattr(p.gender_code, "name", "") if p.gender_code else "",
                    "상태": getattr(p.status_code, "name", "") if p.status_code else "",
                    "투입일": p.deployed_at,
                    "종료일": p.ended_at,
                }
            )
        main_df = pd.DataFrame(rows)
        ref_df = code_reference_df(["person_role", "resident_type", "affiliation", "gender", "person_status"])
        return to_excel_multi_response(
            {"담당자마스터": main_df, "코드참조": ref_df},
            "person_master.xlsx",
            readonly_columns_map={"담당자마스터": ["담당자ID"]},
            code_columns_map={"담당자마스터": ["역할", "상주여부", "소속", "성별", "상태"]},
        )

    if request.method == "POST" and request.POST.get("action") == "import":
        up = request.FILES.get("excel_file")
        if up:
            df = pd.read_excel(up)
            row_refs = []
            duplicate_rows = []
            for _, rec in df.fillna("").iterrows():
                payload = {
                    "employee_no": str(rec.get("사번", "")).strip() or None,
                    "name": str(rec.get("성명", "")).strip(),
                    "role_code": code_from_value("person_role", rec.get("역할")),
                    "resident_type_code": code_from_value("resident_type", rec.get("상주여부")),
                    "affiliation_code": code_from_value("affiliation", rec.get("소속")),
                    "company": str(rec.get("회사명", "")).strip(),
                    "phone": str(rec.get("전화번호", "")).strip(),
                    "email": str(rec.get("내부메일", "")).strip(),
                    "external_email": str(rec.get("외부메일", "")).strip(),
                    "gender_code": code_from_value("gender", rec.get("성별")),
                    "status_code": code_from_value("person_status", rec.get("상태")),
                    "deployed_at": parse_date(rec.get("투입일")),
                    "ended_at": parse_date(rec.get("종료일")),
                }
                person_id = str(rec.get("담당자ID", "")).strip()
                obj = PersonMaster.objects.filter(person_mgmt_no=person_id).first() if person_id else None
                if not obj and payload["employee_no"]:
                    obj = PersonMaster.objects.filter(employee_no=payload["employee_no"]).first()
                if obj:
                    for k, v in payload.items():
                        setattr(obj, k, v)
                    try:
                        obj.save()
                        row_refs.append((obj, {"assigned_service_names": str(rec.get("담당업무", "")).strip()}))
                    except IntegrityError:
                        duplicate_rows.append(payload["employee_no"])
                elif payload["name"]:
                    try:
                        obj = PersonMaster.objects.create(**payload)
                        row_refs.append((obj, {"assigned_service_names": str(rec.get("담당업무", "")).strip()}))
                    except IntegrityError:
                        duplicate_rows.append(payload["employee_no"] or "(사번없음)")
            for person_obj, row in row_refs:
                sync_person_service_assignments_from_row(person_obj, row)
            if duplicate_rows:
                messages.warning(request, f"중복 사번으로 업로드 제외: {', '.join(sorted(set(duplicate_rows)))}")
            messages.success(request, "담당자 마스터 엑셀 업로드 완료")
        return redirect(build_list_redirect(request.path, query=query, page=page))

    if request.method == "POST" and request.POST.get("action") == "save":
        rows = json.loads(request.POST.get("rows_json", "[]"))
        deleted_ids = json.loads(request.POST.get("deleted_ids_json", "[]"))
        PersonMaster.objects.filter(pk__in=deleted_ids).delete()
        row_refs = []
        duplicate_rows = []
        for row in rows:
            pk = str(row.get("id", "")).strip()
            payload = {
                "employee_no": (row.get("employee_no") or "").strip() or None,
                "name": (row.get("name") or "").strip(),
                "role_code": code_from_value("person_role", row.get("role_code")),
                "resident_type_code": code_from_value("resident_type", row.get("resident_type_code")),
                "affiliation_code": code_from_value("affiliation", row.get("affiliation_code")),
                "company": (row.get("company") or "").strip(),
                "phone": (row.get("phone") or "").strip(),
                "email": (row.get("email") or "").strip(),
                "external_email": (row.get("external_email") or "").strip(),
                "gender_code": code_from_value("gender", row.get("gender_code")),
                "status_code": code_from_value("person_status", row.get("status_code")),
                "deployed_at": parse_date(row.get("deployed_at")),
                "ended_at": parse_date(row.get("ended_at")),
            }
            if not payload["name"]:
                continue
            if pk:
                obj = PersonMaster.objects.get(pk=pk)
                exists_other = False
                if payload["employee_no"]:
                    exists_other = PersonMaster.objects.filter(employee_no=payload["employee_no"]).exclude(pk=obj.pk).exists()
                if exists_other:
                    duplicate_rows.append(payload["employee_no"])
                    continue
                for k, v in payload.items():
                    setattr(obj, k, v)
                try:
                    obj.save()
                    row_refs.append((obj, row))
                except IntegrityError:
                    duplicate_rows.append(payload["employee_no"])
            else:
                obj = PersonMaster.objects.filter(employee_no=payload["employee_no"]).first() if payload["employee_no"] else None
                if obj:
                    for k, v in payload.items():
                        setattr(obj, k, v)
                    try:
                        obj.save()
                    except IntegrityError:
                        duplicate_rows.append(payload["employee_no"] or "(사번없음)")
                        continue
                else:
                    try:
                        obj = PersonMaster.objects.create(**payload)
                    except IntegrityError:
                        duplicate_rows.append(payload["employee_no"] or "(사번없음)")
                        continue
                row_refs.append((obj, row))
        for person_obj, row in row_refs:
            sync_person_service_assignments_from_row(person_obj, row)
        if duplicate_rows:
            messages.warning(request, f"중복 사번으로 저장 제외: {', '.join(sorted(set(duplicate_rows)))}")
        messages.success(request, "담당자 마스터 저장 완료")
        return redirect(build_list_redirect(request.path, query=query, page=page))

    page_obj = paginate_all_rows(qs, request.GET.get("page"))
    sa_by_service_role = {}
    for sa in ServiceAttribute.objects.filter(attribute_code_id__in=attribute_codes_for_grid()):
        sa_by_service_role[(sa.service_id, sa.attribute_code_id)] = set(parse_person_ids_from_attr_value(sa.value))

    service_rows = list(ServiceMaster.objects.all().order_by("service_mgmt_no"))
    role_to_acode = {c["role_code"]: c["attribute_code"] for c in SERVICE_PERSON_GRID_COLUMNS}
    for p in page_obj.object_list:
        my_services = []
        my_role = getattr(p.role_code, "code", "") if p.role_code else ""
        my_acode = role_to_acode.get(my_role)
        for svc in service_rows:
            if my_acode and p.pk in sa_by_service_role.get((svc.pk, my_acode), set()):
                my_services.append(svc.name)
        p.assigned_service_names = join_csv_tokens(my_services)

    return render(
        request,
        "web/person_master_list.html",
        {
            "items": page_obj.object_list,
            "page_obj": page_obj,
            "q": query,
            "person_role_options": code_choice_options("person_role"),
            "resident_type_options": code_choice_options("resident_type"),
            "affiliation_options": code_choice_options("affiliation"),
            "gender_options": code_choice_options("gender"),
            "person_status_options": code_choice_options("person_status"),
            "person_status_codes": code_values("person_status"),
            "service_name_options": [s.name for s in service_rows],
        },
    )


@login_required
@permission_required("masters.view_configurationmaster", raise_exception=True)
def configuration_master_list(request):
    query = request.GET.get("q", "").strip()
    page = request.GET.get("page", "").strip()
    qs = (
        ConfigurationMaster.objects.select_related(
            "server_type_code", "operation_dev_code", "infra_type_code", "location_code", "network_zone_code"
        )
        .prefetch_related(
            Prefetch(
                "service_configuration_mappings",
                queryset=ServiceConfigurationMapping.objects.select_related("service"),
            ),
            Prefetch(
                "configuration_component_mappings",
                queryset=ConfigurationComponentMapping.objects.select_related(
                    "component", "component__component_type_code"
                ),
            ),
        )
        .order_by("asset_mgmt_no")
    )
    if query:
        terms, free_text = parse_structured_terms(query)
        if terms:
            for key, value in terms:
                field_key = CONFIG_SEARCH_FIELD_ALIASES.get(key)
                if field_key == CONFIG_CONNECTED_SERVICE_KEY:
                    qs = qs.filter(service_configuration_mappings__service__name__icontains=value)
                elif field_key in CONFIG_CODE_FIELD_MAP:
                    term_q = q_for_code_fields_keyword(value, {field_key: CONFIG_CODE_FIELD_MAP[field_key]})
                    qs = qs.filter(term_q if term_q else Q(pk__in=[]))
                elif field_key in {"hostname", "ip", "url"}:
                    qs = qs.filter(Q(**{f"{field_key}__icontains": value}))
            if free_text:
                free_q = build_model_text_search_q(ConfigurationMaster, free_text)
                free_q |= q_for_code_fields_keyword(free_text, CONFIG_CODE_FIELD_MAP)
                free_q |= Q(service_configuration_mappings__service__name__icontains=free_text)
                qs = qs.filter(free_q)
            qs = qs.distinct()
        else:
            qs = qs.filter(
                build_model_text_search_q(ConfigurationMaster, query)
                | q_for_code_fields_keyword(query, CONFIG_CODE_FIELD_MAP)
                | Q(service_configuration_mappings__service__name__icontains=query)
            ).distinct()

    if request.GET.get("export") == "1":
        component_type_codes_list = code_values("component_type")
        rows = []
        for c in qs:
            labels_by_type: dict[str, list[str]] = {tc: [] for tc in component_type_codes_list}
            for m in c.configuration_component_mappings.all():
                comp = m.component
                if not comp:
                    continue
                tc = getattr(comp.component_type_code, "code", "") or ""
                if tc in labels_by_type:
                    label = " ".join(p for p in (comp.product_name or "", comp.version or "") if p.strip()).strip()
                    labels_by_type[tc].append(label or comp.component_mgmt_no)
            row = {
                "구성ID": c.asset_mgmt_no,
                "구성명": c.hostname,
                "연결서비스": c.connected_services_label,
                "구성유형": getattr(c.server_type_code, "code", ""),
                "운영/개발": getattr(c.operation_dev_code, "code", ""),
                "인프라구분": getattr(c.infra_type_code, "code", ""),
                "위치": getattr(c.location_code, "code", ""),
                "네트웍구분": getattr(c.network_zone_code, "code", ""),
                "IP": c.ip or "",
                "Port": c.port,
                "URL": c.url,
            }
            for tc in component_type_codes_list:
                row[tc] = ", ".join(labels_by_type[tc])
            rows.append(row)
        main_df = pd.DataFrame(rows)
        ref_df = code_reference_df(
            ["config_type", "operation_type", "infra_type", "infra_location", "network_zone", "component_type"]
        )
        code_cols_cfg = ["구성유형", "운영/개발", "인프라구분", "위치", "네트웍구분", *component_type_codes_list]
        return to_excel_multi_response(
            {"구성정보마스터": main_df, "코드참조": ref_df},
            "configuration_master.xlsx",
            readonly_columns_map={"구성정보마스터": ["구성ID"]},
            code_columns_map={"구성정보마스터": code_cols_cfg},
        )

    if request.method == "POST" and request.POST.get("action") == "import":
        up = request.FILES.get("excel_file")
        actor = audit_actor_label(request)
        if up:
            df = pd.read_excel(up)
            component_type_codes_list = code_values("component_type")
            comp_label_to_id: dict[str, int] = {}
            for comp in Component.objects.all():
                label = " ".join(p for p in (comp.product_name or "", comp.version or "") if p.strip()).strip()
                if label and label not in comp_label_to_id:
                    comp_label_to_id[label] = comp.pk
            for _, rec in df.fillna("").iterrows():
                # 네트웍 구분 컬럼은 신규 라벨 우선, 구 라벨("네트워크")도 호환 유지
                nz_raw = rec.get("네트웍구분") if "네트웍구분" in rec.index else rec.get("네트워크", "")
                payload = {
                    "hostname": str(rec.get("구성명", "")).strip(),
                    "server_type_code": code_from_value("config_type", rec.get("구성유형")),
                    "operation_dev_code": code_from_value("operation_type", rec.get("운영/개발")),
                    "infra_type_code": code_from_value("infra_type", rec.get("인프라구분")),
                    "location_code": code_from_value("infra_location", rec.get("위치")),
                    "network_zone_code": code_from_value("network_zone", nz_raw),
                    "ip": str(rec.get("IP", "")).strip() or None,
                    "port": str(rec.get("Port", "")).strip(),
                    "url": str(rec.get("URL", "")).strip(),
                }
                cfg_id = str(rec.get("구성ID", "")).strip()
                obj = ConfigurationMaster.objects.filter(asset_mgmt_no=cfg_id).first() if cfg_id else None
                if obj:
                    for k, v in payload.items():
                        setattr(obj, k, v)
                    obj.updated_by = actor
                    obj.save()
                    sync_service_configuration_mappings(obj, str(rec.get("연결서비스", "")))
                elif payload["hostname"]:
                    obj = ConfigurationMaster.objects.create(**payload, created_by=actor, updated_by=actor)
                    sync_service_configuration_mappings(obj, str(rec.get("연결서비스", "")))
                else:
                    obj = None
                if obj is not None:
                    desired_ids: list[int] = []
                    for tc in component_type_codes_list:
                        raw_cell = str(rec.get(tc, "")).strip()
                        if not raw_cell:
                            continue
                        for token in [t.strip() for t in raw_cell.split(",") if t.strip()]:
                            cid = comp_label_to_id.get(token)
                            if cid is not None:
                                desired_ids.append(cid)
                    seen: set[int] = set()
                    ConfigurationComponentMapping.objects.filter(configuration=obj).delete()
                    for cid in desired_ids:
                        if cid in seen:
                            continue
                        seen.add(cid)
                        ConfigurationComponentMapping.objects.create(configuration=obj, component_id=cid)
            messages.success(request, "구성정보 마스터 엑셀 업로드 완료")
        return redirect(build_list_redirect(request.path, query=query, page=page))

    if request.method == "POST" and request.POST.get("action") == "save":
        rows = json.loads(request.POST.get("rows_json", "[]"))
        deleted_ids = json.loads(request.POST.get("deleted_ids_json", "[]"))
        actor = audit_actor_label(request)
        component_type_codes_list = code_values("component_type")
        with transaction.atomic():
            ConfigurationMaster.objects.filter(pk__in=deleted_ids).delete()
            for row in rows:
                pk = str(row.get("id", "")).strip()
                connected_raw = row.get("connected_service") or ""
                payload = {
                    "hostname": (row.get("hostname") or "").strip(),
                    "server_type_code": code_from_value("config_type", row.get("server_type_code")),
                    "operation_dev_code": code_from_value("operation_type", row.get("operation_dev_code")),
                    "infra_type_code": code_from_value("infra_type", row.get("infra_type_code")),
                    "location_code": code_from_value("infra_location", row.get("location_code")),
                    "network_zone_code": code_from_value("network_zone", row.get("network_zone_code")),
                    "ip": (row.get("ip") or "").strip() or None,
                    "port": (row.get("port") or "").strip(),
                    "url": (row.get("url") or "").strip(),
                }
                obj = None
                if pk:
                    obj = ConfigurationMaster.objects.get(pk=pk)
                    for k, v in payload.items():
                        setattr(obj, k, v)
                    obj.updated_by = actor
                    obj.save()
                    sync_service_configuration_mappings(obj, connected_raw)
                elif payload["hostname"]:
                    obj = ConfigurationMaster.objects.create(**payload, created_by=actor, updated_by=actor)
                    sync_service_configuration_mappings(obj, connected_raw)
                if obj:
                    sync_configuration_component_mappings(obj, row, component_type_codes_list)
        messages.success(request, "구성정보 마스터 저장 완료")
        return redirect(build_list_redirect(request.path, query=query, page=page))

    page_obj = paginate_all_rows(qs, request.GET.get("page"))
    service_name_options = sorted(
        {n for n in ServiceMaster.objects.values_list("name", flat=True) if (n or "").strip()}
    )

    component_type_options = code_choice_options("component_type")
    component_type_codes_list = [opt["code"] for opt in component_type_options]
    active_type_codes = set(component_type_codes_list)

    components_all: list[dict] = []
    for comp in (
        Component.objects.select_related("component_type_code")
        .order_by("product_name", "version", "component_mgmt_no")
    ):
        type_code = getattr(comp.component_type_code, "code", "") or ""
        if type_code not in active_type_codes:
            continue
        label = " ".join(p for p in (comp.product_name or "", comp.version or "") if p.strip()).strip()
        components_all.append(
            {
                "id": comp.pk,
                "product_name": comp.product_name or "",
                "version": comp.version or "",
                "vendor_name": comp.vendor_name or "",
                "type": type_code,
                "label": label or comp.component_mgmt_no,
            }
        )

    aliases_by_product: dict[str, list[str]] = {}
    for product_name, alias in ComponentAlias.objects.values_list("product_name", "alias").order_by("product_name", "alias"):
        aliases_by_product.setdefault(product_name, []).append(alias)

    for c in page_obj.object_list:
        ids_by_type: dict[str, list[str]] = {tc: [] for tc in component_type_codes_list}
        labels_by_type: dict[str, list[str]] = {tc: [] for tc in component_type_codes_list}
        for mapping in c.configuration_component_mappings.all():
            comp = mapping.component
            if not comp:
                continue
            type_code = getattr(comp.component_type_code, "code", "") or ""
            if type_code not in ids_by_type:
                continue
            ids_by_type[type_code].append(str(comp.pk))
            label = " ".join(p for p in (comp.product_name or "", comp.version or "") if p.strip()).strip()
            labels_by_type[type_code].append(label or comp.component_mgmt_no)
        c.component_columns = [
            {
                "type": tc,
                "ids": ",".join(ids_by_type[tc]),
                "labels": ", ".join(labels_by_type[tc]),
            }
            for tc in component_type_codes_list
        ]

    return render(
        request,
        "web/configuration_master_list.html",
        {
            "items": page_obj.object_list,
            "page_obj": page_obj,
            "q": query,
            "config_type_codes": code_values("config_type"),
            "operation_type_codes": code_values("operation_type"),
            "infra_type_codes": code_values("infra_type"),
            "infra_location_codes": code_values("infra_location"),
            "network_zone_codes": code_values("network_zone"),
            "service_name_options": service_name_options,
            "component_type_options": component_type_options,
            "components_all_json": json.dumps(components_all, ensure_ascii=False),
            "component_type_codes_json": json.dumps(component_type_codes_list, ensure_ascii=False),
            "aliases_by_product_json": json.dumps(aliases_by_product, ensure_ascii=False),
        },
    )


@login_required
@permission_required("masters.view_component", raise_exception=True)
def component_master_list(request):
    query = request.GET.get("q", "").strip()
    page = request.GET.get("page", "").strip()
    qs = Component.objects.select_related("component_type_code", "support_status_code").order_by("component_mgmt_no")
    if query:
        terms, free_text = parse_structured_terms(query)
        if terms:
            for key, value in terms:
                field_key = COMPONENT_SEARCH_FIELD_ALIASES.get(key)
                if field_key in COMPONENT_CODE_FIELD_MAP:
                    term_q = q_for_code_fields_keyword(value, {field_key: COMPONENT_CODE_FIELD_MAP[field_key]})
                    qs = qs.filter(term_q if term_q else Q(pk__in=[]))
                elif field_key in {"product_name", "version", "vendor_name", "cpe_name"}:
                    qs = qs.filter(Q(**{f"{field_key}__icontains": value}))
            if free_text:
                free_q = build_model_text_search_q(Component, free_text)
                free_q |= q_for_code_fields_keyword(free_text, COMPONENT_CODE_FIELD_MAP)
                qs = qs.filter(free_q)
            qs = qs.distinct()
        else:
            qs = qs.filter(build_model_text_search_q(Component, query) | q_for_code_fields_keyword(query, COMPONENT_CODE_FIELD_MAP)).distinct()

    if request.GET.get("export") == "1":
        rows = [
            {
                "컴포넌트ID": c.component_mgmt_no,
                "컴포넌트명": c.product_name,
                "버전": c.version,
                "유형": getattr(c.component_type_code, "name", "") if c.component_type_code else "",
                "벤더명": c.vendor_name,
                "CPE": c.cpe_name,
                "EOS": c.eos_date,
                "EOL": c.eol_date,
                "지원여부": getattr(c.support_status_code, "code", "") if c.support_status_code else "",
            }
            for c in qs
        ]
        main_df = pd.DataFrame(rows)
        ref_df = code_reference_df(["component_type", "support_status"])
        return to_excel_multi_response(
            {"컴포넌트마스터": main_df, "코드참조": ref_df},
            "component_master.xlsx",
            code_columns_map={"컴포넌트마스터": ["유형", "지원여부"]},
        )

    if request.method == "POST" and request.POST.get("action") == "import":
        up = request.FILES.get("excel_file")
        actor = audit_actor_label(request)
        if up:
            df = pd.read_excel(up)
            for _, rec in df.fillna("").iterrows():
                payload = {
                    "product_name": str(rec.get("컴포넌트명", "")).strip(),
                    "version": str(rec.get("버전", "")).strip(),
                    "component_type_code": code_from_value("component_type", rec.get("유형")),
                    "vendor_name": str(rec.get("벤더명", "")).strip(),
                    "cpe_name": str(rec.get("CPE", "")).strip(),
                    "eos_date": parse_date(rec.get("EOS")),
                    "eol_date": parse_date(rec.get("EOL")),
                    "support_status_code": code_from_value("support_status", rec.get("지원여부")),
                }
                cmp_id = str(rec.get("컴포넌트ID", "")).strip()
                obj = Component.objects.filter(component_mgmt_no=cmp_id).first() if cmp_id else None
                if obj:
                    for k, v in payload.items():
                        setattr(obj, k, v)
                    obj.updated_by = actor
                    obj.save()
                elif payload["product_name"]:
                    Component.objects.create(**payload, created_by=actor, updated_by=actor)
            messages.success(request, "컴포넌트 마스터 엑셀 업로드 완료")
        return redirect(build_list_redirect(request.path, query=query, page=page))

    if request.method == "POST" and request.POST.get("action") == "save":
        rows = json.loads(request.POST.get("rows_json", "[]"))
        deleted_ids = json.loads(request.POST.get("deleted_ids_json", "[]"))
        actor = audit_actor_label(request)
        Component.objects.filter(pk__in=deleted_ids).delete()
        for row in rows:
            pk = str(row.get("id", "")).strip()
            payload = {
                "component_type_code": code_from_value("component_type", row.get("component_type_code")),
                "product_name": (row.get("product_name") or "").strip(),
                "version": (row.get("version") or "").strip(),
                "vendor_name": (row.get("vendor_name") or "").strip(),
                "cpe_name": (row.get("cpe_name") or "").strip(),
                "eos_date": parse_date(row.get("eos_date")),
                "eol_date": parse_date(row.get("eol_date")),
                "support_status_code": code_from_value("support_status", row.get("support_status_code")),
            }
            if pk:
                obj = Component.objects.get(pk=pk)
                for k, v in payload.items():
                    setattr(obj, k, v)
                obj.updated_by = actor
                obj.save()
            elif payload["product_name"]:
                Component.objects.create(**payload, created_by=actor, updated_by=actor)
        messages.success(request, "컴포넌트 마스터 저장 완료")
        return redirect(build_list_redirect(request.path, query=query, page=page))

    page_obj = paginate_all_rows(qs, request.GET.get("page"))
    return render(
        request,
        "web/component_master_list.html",
        {
            "items": page_obj.object_list,
            "page_obj": page_obj,
            "q": query,
            "component_type_options": code_choice_options("component_type"),
            "support_status_codes": code_values("support_status"),
        },
    )


@login_required
def ai_asset_search(request):
    query = request.GET.get("q", "").strip()
    results = []
    error = ""
    if query:
        try:
            svc = AssetSearchService()
            results = svc.search(query, k=10)
        except FileNotFoundError:
            error = "AI 인덱스가 없습니다. `python manage.py build_asset_index`를 먼저 실행하세요."
    return render(request, "web/ai_asset_search.html", {"q": query, "results": results, "error": error})
